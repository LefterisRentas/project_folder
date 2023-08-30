import openpyxl

def process_excel(excel_file):
    # Open the Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    # Iterate through rows and implement logic for filtering orders and calculating totals
    for row in sheet.iter_rows(min_row=7):  # Assuming data starts from row 7
            #Implement your logic here
            pass
    # Create and save the output Excel file
    output_workbook = openpyxl.Workbook()
    output_sheet = output_workbook.active
    # Implement logic to write filtered data to the output sheet
    output_workbook.save("output.xlsx")

def main():
    excel_file = "path_to_excel.xlsx"  # Path to the selected Excel file
    exclude_file = "path_to_exclude.txt"  # Path to the selected exclude file
    amount = 125  # Amount input by the user
    process_excel(excel_file)

if __name__ == "__main__":
    main()
