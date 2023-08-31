import os
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
import threading
import subprocess
import tkinter.ttk as ttk


class EntryField(ttk.Entry):
    def __init__(self, master=None, **kwargs):
        super().__init__(master, **kwargs)
        self.config(font=("Helvetica", 12))


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Order Processor")
        self.root.geometry("500x800")

        self.style = ttk.Style()
        self.style.theme_use("clam")

        # Configure label style
        self.style.configure("TLabel",
                             foreground="#333",
                             font=("Helvetica", 14, "bold"))

        # Configure button style
        self.style.configure("TButton",
                             background="#2196F3",  # Blue background color
                             foreground="white",
                             padding=10,
                             font=("Helvetica", 12, "bold"))

        # Configure button style on hover
        self.style.map("TButton",
                       background=[("active", "#1976D2")],  # Darker blue on hover
                       foreground=[("active", "white")])

        self.excel_file_label = ttk.Label(root, text="Excel File:")
        self.excel_file_label.pack(pady=10)

        self.excel_file_entry = ttk.Entry(root)
        self.excel_file_entry.pack()

        self.excel_file_button = ttk.Button(root, text="Browse Excel", command=self.browse_excel)
        self.excel_file_button.pack()

        # Button to generate exclude.txt
        self.generate_exclude_button = ttk.Button(root, text="Generate Exclude.txt", command=self.generate_exclude_file)
        self.generate_exclude_button.pack()

        self.exclude_file_label = ttk.Label(root, text="Exclude File:")
        self.exclude_file_label.pack()

        self.exclude_file_entry = ttk.Entry(root)
        self.exclude_file_entry.pack()

        self.exclude_file_button = ttk.Button(root, text="Browse Exclude", command=self.browse_exclude)
        self.exclude_file_button.pack()

        self.exclude_file_label = ttk.Label(root, text="Exclude text:")
        self.exclude_file_label.pack()

        # Entry field for adding exclude strings
        self.exclude_entry = ttk.Entry(root)
        self.exclude_entry.pack()

        # Button to add exclude strings
        self.add_exclude_button = ttk.Button(root, text="Add Exclude", command=self.add_exclude_string)
        self.add_exclude_button.pack(pady=5)

        # Button to remove selected exclude
        self.remove_exclude_button = ttk.Button(root, text="Remove Selected", command=self.remove_selected_exclude)
        self.remove_exclude_button.pack(pady=5)

        self.exclude_listbox = tk.Listbox(root)
        self.exclude_listbox.pack(pady=10)

        self.save_changes_button = ttk.Button(root, text="Save Changes", command=self.save_changes)
        self.save_changes_button.pack(pady=10)

        self.amount_label = ttk.Label(root, text="Amount:")
        self.amount_label.pack()

        self.amount_entry = ttk.Entry(root)
        self.amount_entry.pack()

        self.process_button = ttk.Button(root, text="Process", command=self.process_files)
        self.process_button.pack()

        # List to store exclude strings
        self.exclude_strings = []

    def browse_excel(self):
        excel_file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
        self.excel_file_entry.delete(0, tk.END)
        self.excel_file_entry.insert(0, excel_file)

    def browse_exclude(self):
        exclude_file = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if exclude_file:
            with open(exclude_file, "r", encoding="utf-8") as f:
                self.exclude_strings = [line.strip() for line in f.readlines()]

            self.exclude_listbox.delete(0, tk.END)  # Clear the listbox
            for exclude_string in self.exclude_strings:
                self.exclude_listbox.insert(tk.END, exclude_string)

            self.exclude_file_entry.delete(0, tk.END)
            self.exclude_file_entry.insert(0, exclude_file)

    def add_exclude_string(self):
        exclude_string = self.exclude_entry.get()
        if exclude_string:
            if exclude_string not in self.exclude_strings:
                self.exclude_strings.append(exclude_string)
                self.exclude_listbox.insert(tk.END, exclude_string)
                self.exclude_entry.delete(0, tk.END)
            else:
                messagebox.showinfo("Duplicate Entry", "The exclude string already exists.")

    def remove_selected_exclude(self):
        selected_index = self.exclude_listbox.curselection()
        if selected_index:
            selected_index = selected_index[0]
            selected_exclude = self.exclude_listbox.get(selected_index)
            confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete '{selected_exclude}'?")
            if confirm:
                self.exclude_listbox.delete(selected_index)
                self.exclude_strings.remove(selected_exclude)

    def generate_exclude_file(self):
        exclude_file_path = "exclude.txt"

        # Write exclude strings to the file with UTF-8 encoding
        with open(exclude_file_path, "w", encoding="utf-8") as f:
            for exclude_string in self.exclude_strings:
                f.write(exclude_string + "\n")
        # Update the exclude_file_entry
        self.exclude_file_entry.delete(0, tk.END)
        self.exclude_file_entry.insert(0, os.path.realpath(f.name))

        # Clear the exclude_strings list and the exclude_listbox
        self.exclude_strings.clear()  # Clear the list
        self.exclude_listbox.delete(0, tk.END)  # Clear the listbox
        return f.name

    def save_changes(self):
        # Update the displayed excluded strings in the listbox
        self.exclude_listbox.delete(0, tk.END)
        for exclude_string in self.exclude_strings:
            self.exclude_listbox.insert(tk.END, exclude_string)
        file = self.generate_exclude_file()  # Save changes to file

        # Update the listbox with the latest exclude strings
        self.exclude_listbox.delete(0, tk.END)
        if file:
            with open(file, "r", encoding="utf-8") as f:
                self.exclude_strings = [line.strip() for line in f.readlines()]

            self.exclude_listbox.delete(0, tk.END)  # Clear the listbox
            for exclude_string in self.exclude_strings:
                self.exclude_listbox.insert(tk.END, exclude_string)

            self.exclude_file_entry.delete(0, tk.END)
            self.exclude_file_entry.insert(0, os.path.abspath(file))

    def process_files(self):
        excel_file = self.excel_file_entry.get()
        exclude_file = self.exclude_file_entry.get()
        amount = float(self.amount_entry.get())

        # Start a new thread for processing
        processing_thread = threading.Thread(target=self.process_excel,
                                             args=(excel_file, exclude_file, amount, self.exclude_strings))
        processing_thread.start()

    @staticmethod
    def process_excel(excel_file, exclude_file, amount, exclude_strings):
        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Identify column indices based on headers
        header_row = 6
        column_mapping = {}
        for col_idx, cell in enumerate(sheet[header_row], start=1):
            header = cell.value
            if header == "Συνολική":
                column_mapping["total"] = col_idx
            elif header == "Διεύθυνση":
                column_mapping["address"] = col_idx
            elif header == "Επωνυμία":
                column_mapping["name"] = col_idx

        # Calculate totals and exclude orders
        address_totals = {}  # Nested dictionary: address -> name -> total
        for row in sheet.iter_rows(min_row=header_row + 1):
            total_value = row[column_mapping["total"] - 1].value
            address = row[column_mapping["address"] - 1].value
            name = row[column_mapping["name"] - 1].value

            if total_value is not None and address is not None and name is not None:
                if name not in exclude_strings:
                    address_totals.setdefault(address, {})
                    address_totals[address].setdefault(name, 0)
                    address_totals[address][name] += total_value

        output_data = []
        for address, name_totals in address_totals.items():
            for name, total in name_totals.items():
                if total <= amount:
                    output_data.append((name, address, total))

        # Create and save the output Excel file
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active
        output_sheet.append(["Επωνυμία", "Διεύθυνση", "Συνολική"])
        for entry in output_data:
            output_sheet.append(entry)
        output_workbook.save("output.xlsx")

        # Open the output.xlsx file with its default application
        subprocess.call(["start", "output.xlsx"], shell=True)


if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
